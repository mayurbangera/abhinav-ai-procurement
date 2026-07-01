from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font


class ExcelService:

    @staticmethod
    def export_suppliers(suppliers):

        workbook = Workbook()

        sheet = workbook.active

        sheet.title = "Suppliers"

        headers = [
            "Supplier ID",
            "Company Name",
            "Principal Business",
            "GST Number",
            "Company Address",
            "Contact Person",
            "Email Address",
            "WhatsApp Number",
            "Supplier Category",
            "Material Types",
            "Bank Name",
            "Account Holder Name",
            "Account Number",
            "IFSC Code",
            "Branch Name",
            "MSME Registered",
            "MSME Number",
            "Customer Reference",
            "Authorized Person",
            "Designation",
            "Declaration Accepted",
            "Registration Status",
            "Created At"
        ]

        # Write Header
        for column, header in enumerate(headers, start=1):

            cell = sheet.cell(
                row=1,
                column=column
            )

            cell.value = header
            cell.font = Font(bold=True)

        # Write Data
        row = 2

        for supplier in suppliers:

            sheet.cell(row=row, column=1).value = supplier.id
            sheet.cell(row=row, column=2).value = supplier.company_name
            sheet.cell(row=row, column=3).value = supplier.principal_business
            sheet.cell(row=row, column=4).value = supplier.gst_number
            sheet.cell(row=row, column=5).value = supplier.registered_address
            sheet.cell(row=row, column=6).value = supplier.contact_person_name
            sheet.cell(row=row, column=7).value = supplier.contact_person_email
            sheet.cell(row=row, column=8).value = supplier.whatsapp_number
            sheet.cell(row=row, column=9).value = supplier.supplier_category
            sheet.cell(row=row, column=10).value = supplier.material_types
            sheet.cell(row=row, column=11).value = supplier.bank_name
            sheet.cell(row=row, column=12).value = supplier.beneficiary_name
            sheet.cell(row=row, column=13).value = supplier.bank_account_number
            sheet.cell(row=row, column=14).value = supplier.bank_ifsc
            sheet.cell(row=row, column=15).value = supplier.branch_name

            sheet.cell(row=row, column=16).value = (
                "YES" if supplier.is_msme else "NO"
            )

            sheet.cell(row=row, column=17).value = supplier.msme_number
            sheet.cell(row=row, column=18).value = supplier.references
            sheet.cell(row=row, column=19).value = supplier.authorized_person_name
            sheet.cell(row=row, column=20).value = supplier.designation

            sheet.cell(row=row, column=21).value = (
                "YES" if supplier.declaration_accepted else "NO"
            )

            sheet.cell(row=row, column=22).value = supplier.registration_status
            sheet.cell(row=row, column=23).value = str(supplier.created_at)

            row += 1

        excel_file = BytesIO()

        workbook.save(excel_file)

        excel_file.seek(0)

        return excel_file